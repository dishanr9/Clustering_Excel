import getpass

import pandas as pd
import numpy as np
import traceback
import natsort

import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
import seaborn as sns
import openpyxl

import string
import sys

import xlsxwriter

import view
import os
import math

from numpy import pi, log2, e
from collections import Counter, OrderedDict
from sklearn.mixture import GaussianMixture as GMM
from scipy.spatial import distance
from sklearn.metrics.cluster import adjusted_rand_score
from sklearn.metrics import rand_score

from openpyxl.styles import Border, Side, Alignment

# Global variables
covariances = {}
labels = {}
means = {}
entropy = {}
record_count = {}
label_counts = {}
full_data = {}
id_labels = {}
cluster_instants = OrderedDict()

# Will always contain only 2 keys - 0 and 1
means_ = {}
covariances_ = {}
entropy_ = {}
merged_centerset = {}
clusters = {}
# rows = 0
columns = 0
notations = {}
# Will always contain only 2 keys - 0 and 1
datapoints = {0: {}, 1: {}}

# original model
core_model = {}
# Merged model
merged_model = {}
# Consolidation model
consol_model = {0: {}, 1: {}}

# evolving distance
evolving_distance = {}
char_list = list(string.ascii_uppercase) + list(string.ascii_lowercase)
center_set = {}
cons_list = ['A' + i for i in list(string.ascii_uppercase)] + ['B' + i for i in list(string.ascii_uppercase)] + ['C' + i
                                                                                                                 for i
                                                                                                                 in
                                                                                                                 list(
                                                                                                                     string.ascii_uppercase)] + [
                'D' + i for i in list(string.ascii_uppercase)] + ['E' + i for i in list(string.ascii_uppercase)] + [
                'F' + i for i in list(string.ascii_uppercase)]

n = 60
chunk_size = 500
noise = pd.DataFrame({})

dataset = "A3.xlsx"
prefix = dataset[0:-5]
folder_name = 'C:/Users/disha/Desktop/HPBAckUp/Thesis/Realtime_Datastream_Clustering/'+prefix+'/'
merge_file = folder_name + prefix + "_" + str(chunk_size) + "_DC_1.xlsx"
table_file = folder_name + prefix + "_" + str(chunk_size) + "_df_entropy.xlsx"
merge_ntop = folder_name + prefix + "_" + str(chunk_size) + "_df_ntop.xlsx"
cons_entropy = folder_name + prefix + "_" + str(chunk_size) + '_Consolidate_entropy_1.xlsx'
cons_ntop = folder_name + prefix + "_" + str(chunk_size) + '_df_ntop_cons_2base.xlsx'
tt_file = folder_name + prefix + "_" + str(chunk_size) + '_time_trace.xlsx'


# Excel reading happening thrice
# Remove console.logs
# Block out holding data as Y. if not for time trace

def set_limits(excel=dataset):
    if excel:
        data = pd.read_excel(excel, skiprows=14, usecols='B:D', names=["x", "y","truth"], engine='openpyxl')
        x_lim, y_lim = 10 ** math.ceil(math.log10(data['x'].max())), 10 ** math.ceil(math.log10(data['y'].max()))

        # return (0, x_lim), (0, y_lim)

        def lim(max_num):
            exp = len(str(max_num)) - 1
            #     exp = math.floor(math.log10(n))
            q = math.floor(max_num / 10 ** exp)
            lim = (q + 1) * (10 ** exp)
            return lim

    x_lim, y_lim = lim(data['x'].max()), lim(data['y'].max())
    return (0, x_lim), (0, y_lim)


(xmin, xmax), (ymin, ymax) = set_limits(dataset)
print("limits:", (xmin, xmax), (ymin, ymax))


def add_noise(max_x, max_y, extra=7500):
    x = np.random.randint(0, max_x, extra)
    y = np.random.randint(0, max_y, extra)
    noise = pd.DataFrame({'x': x, 'y': y})
    return noise


def duplicate_dataset(df):
    df = df.append(df, ignore_index =True)
    return df


def gmm_fulldata(excel=dataset):
    global noise
    if excel:
        incoming_data = pd.read_excel(excel, skiprows=14, usecols='B:D', names=["x", "y","truth"], engine='openpyxl')
        incoming_data = incoming_data.sample(frac=1).reset_index(drop=True)
        # if noise.empty:
        #     noise = add_noise(max(incoming_data['x']), max(incoming_data['y']))

        # df = incoming_data.copy()
        # df = df + 250
        # incoming_data = incoming_data.append(noise, ignore_index=True)
        # incoming_data = duplicate_dataset(incoming_data)
        # when gmm is run with n=number of clusters we end up with in our algorithm
        clusters = [len(set(consol_model[1]['data']['id_label'])), 50]

        for i in range(len(clusters)):
            g = GMM(n_components=clusters[i])
            g = g.fit(incoming_data[["x", "y"]])
            l = g.predict(incoming_data[["x", "y"]])
            w_mean = g.means_
            label = 't_label_' + str(i)
            incoming_data[label] = ["w" + "_" + str(x) for x in l]

            lbl_count = dict(Counter(incoming_data[label]))
            keys = natsort.natsorted(lbl_count.keys())
            l_dict = OrderedDict((k, lbl_count.get(k)) for k in keys)

            # sns.set(rc={'figure.figsize': (10, 6)})
            # n_cols = {23: 1, 46: 2, 92: 3, 184: 4}
            # n_col = 1
            # right = 0.7
            # for m in n_cols.keys():
            #     if len(set(incoming_data[label])) <= m:
            #         n_col = n_cols.get(m)
            #         if n_col > 2:
            #             right = 0.65
            #         break
            # plt.clf()
            # fig, ax = plt.subplots(figsize=(10, 8))
            # scatter_plot = sns.scatterplot(data=incoming_data, x="x", y="y", hue=label, s=15,
            #                                palette=view.generate_colors(len(set(incoming_data[label]))))
            # plt.legend(bbox_to_anchor=(1.04, 1), borderaxespad=0, ncol=n_col)
            # plt.setp(scatter_plot.get_legend().get_texts(), fontsize='12')
            # # plt.scatter(gmm.means_[:, 0], gmm.means_[:, 1], s=60, color="black", marker="+")
            # plt.xlim(xmin, xmax)
            # plt.ylim(ymin, ymax)
            # s = "At once GMM, variable n=" + str(clusters[i]) + "chunk_size=" + str(chunk_size)
            # if i:
            #     s = "At once fixed n=" + str(clusters[i]) + "chunk_size=" + str(chunk_size)
            # plt.title(s)
            #
            # for m in range(clusters[i]):
            #     plt.text(g.means_[m][0], g.means_[m][1], keys[m],
            #              horizontalalignment='center', verticalalignment='center', size=8, weight='bold',
            #              color='black', backgroundcolor='white')
            #
            # plt.tight_layout(rect=[0, 0, 0.85, 1])
            #
            # pts = str(len(incoming_data[label]))
            # plt.figtext(0.5, 0.01, "no. of points={0}".format(pts), ha="center", fontsize=10,
            #             bbox={"facecolor": "orange", "alpha": 0.5, "pad": 5})
            #
            # # plt.subplots_adjust(right=0.7)
            # name = "full_data_gmm_" + str(i) + ".png"
            # plt.savefig(name, bbox_inches="tight")
            # plt.close()


        return incoming_data
    else:
        return None


def read_data(excel=dataset):
    global merge_file, table_file, merge_ntop, cons_entropy, cons_ntop, noise
    if excel:
        incoming_data = pd.read_excel(excel, skiprows=14, usecols='B:D', names=["x", "y","truth"], engine='openpyxl')
        incoming_data = incoming_data.sample(frac=1).reset_index(drop=True)
        # if noise.empty:
        #     noise = add_noise(max(incoming_data['x']), max(incoming_data['y']))
        # df = incoming_data.copy()
        # df = df + 250
        # incoming_data = incoming_data.append(noise, ignore_index=True)
        incoming_data = duplicate_dataset(incoming_data)

        print("Len of incoming data: ", len(incoming_data))

        data_chunks = np.array_split(incoming_data, int(len(incoming_data) / chunk_size))
        print("data_chunks:", len(data_chunks))

        files = [merge_file, table_file, merge_ntop, cons_entropy, cons_ntop, tt_file]

        for file in files:
            if os.path.exists(os.getcwd() + '\\' + file):
                os.remove(file)
            w = xlsxwriter.Workbook(file)
            worksheet = w.add_worksheet()
            w.close()

        # merge_file = prefix + merge_file
        # table_file = prefix + table_file
        # merge_ntop = prefix + merge_ntop
        # cons_entropy = prefix + cons_entropy
        # cons_ntop = prefix + cons_ntop

        return data_chunks
    else:
        raise FileNotFoundError("File does not exist!")


def clustering(data_chunks, i):
    distances = []
    cmap = plt.cm.jet
    cmap = mcolors.ListedColormap(plt.rcParams['axes.prop_cycle'].by_key()['color'])

    norm = mcolors.BoundaryNorm(np.arange(n + 1) - 0.5, n)

    d = data_chunks[["x", "y","truth"]]

    gmm = GMM(n_components=n, tol=0)
    gmm = gmm.fit(d)

    record_count[i] = len(d)
    labels[i] = gmm.predict(d)
    means[i] = gmm.means_
    covariances[i] = gmm.covariances_

    # Calculating entropy using covariance matrices
    entropy[i] = 0.5 * (len(covariances[i]) * log2(2 * pi * e) + log2(abs(np.linalg.det(covariances[i]))))

    X = data_chunks

    Y = X
    Y["labels"] = labels[i]

    # Y["label_2"] = [(10 * i) + label for label in labels[i]]

    X = X.values

    centers = np.column_stack((gmm.means_[:, 0], gmm.means_[:, 1]))
    # Add centers series to center_set
    center_set[i] = centers

    print("centers: \n", centers)

    id_label = ["D" + str(i) + "_" + str(x) for x in labels[i]]
    print("id_label \n", id_label)
    Y['id_label'] = id_label
    id_labels[i] = id_label

    # Hold each data_chunk
    full_data[i] = Y
    print("Check : ", dict(Counter(id_labels[i])))
    if i == 2:
        t_mean = {}
        for k, v in dict(Counter(id_labels[i])).items():
            t_df = full_data[i][full_data[i]['id_label'] == k][["x", "y","truth"]]
            t_mean[k] = np.mean(t_df, axis=0)

        pd.DataFrame(t_mean).to_excel("t_mean.xlsx")

    # Get a count of points in each label
    label_counts[i] = dict(Counter(id_labels[i]))
    keys = natsort.natsorted(label_counts[i].keys())
    l = OrderedDict((k, label_counts[i].get(k)) for k in keys)
    label_counts[i] = l
    #
    # facet = sns.lmplot(data=Y, x='x', y='y', hue='id_label',
    #                    fit_reg=False, legend=True, legend_out=True)

    # plt.show()

    a = center_set[i]
    a = np.array(a)
    b = covariances[i]
    b = np.array(b)
    c = np.zeros((n,), dtype=np.object)
    # cov = np.zeros((n,) ,dtype=np.object)
    cov = [0] * n
    for o in range(n):
        c[o] = a[o, :]
        cov[o] = b[o, :, :]

    core_model['model_' + str(i)] = pd.DataFrame({
        "Cluster": list(label_counts[i].keys()),
        "No_of_pts": list(label_counts[i].values()),
        "Entropy": entropy[i],
        "Mean": c,
        "Covariance": cov,

    })

    data_1 = core_model['model_' + str(i)]
# """
#     sns.set(rc={'figure.figsize': (10, 6)})
#     n_cols = {23: 1, 46: 2, 92: 3, 184: 4}
#     n_col = 1
#     right = 0.7
#     for m in n_cols.keys():
#         if len(set(Y['id_label'])) <= m:
#             n_col = n_cols.get(m)
#             if n_col > 2:
#                 right = 0.65
#             break
#     plt.clf()
#     fig, ax = plt.subplots(figsize=(10, 6))
#     g = sns.scatterplot(data=Y, x="x", y="y", hue="id_label", s=15,
#                         palette=view.generate_colors(len(set(id_label))))
#     plt.legend(bbox_to_anchor=(1.04, 1), borderaxespad=0, ncol=n_col)
#     plt.setp(g.get_legend().get_texts(), fontsize='12')
#     s = "No. of clusters = " + str(len(list(label_counts[i].keys())))
#     plt.title(s)
#
#     for m in range(len(data_1)):
#         plt.text(data_1.loc[m, "Mean"][0], data_1.loc[m, "Mean"][1], data_1.loc[m, "Cluster"],
#                  horizontalalignment='center', verticalalignment='center', size=8, weight='bold',
#                  color='black', backgroundcolor='white')
#
#     plt.xlim(xmin, xmax)
#     plt.ylim(ymin, ymax)
#
#     plt.tight_layout(rect=[0, 0, 0.75, 1])
#     plt.figtext(0.5, 0.01, "no. of points=" + str(data_1['No_of_pts'].sum()), ha="center", fontsize=10,
#                 bbox={"facecolor": "orange", "alpha": 0.5, "pad": 5})
#     name = 'Chunk_' + str(i) + ".png"
#     plt.subplots_adjust(right=right)
#     plt.savefig(name, bbox_inches="tight")
#     plt.close()
#     """

    k = 0 if i == 0 else 1
    merged_model[k] = core_model['model_' + str(i)].copy()

    datapoints[k]['data'] = Y
    datapoints[k]['image'] = "i"

    f = dataset[0:-4] + "_ m0_" + str(i) + ".xlsx"
    w = xlsxwriter.Workbook(f)
    worksheet = w.add_worksheet()

    merged_model[0] = merged_model[0].sort_values(by="Cluster", key=natsort.natsort_keygen(), ignore_index=True)
    # """
    # merged_model[0].to_excel(f)
    # """

    if 1 in merged_model:
# """
#         f2 = dataset[0:-4] + "_ m1_" + str(i) + ".xlsx"
#         w = xlsxwriter.Workbook(f2)
#         worksheet = w.add_worksheet()
# """

        merged_model[1] = merged_model[1].sort_values(by="Cluster", key=natsort.natsort_keygen(), ignore_index=True)
# """
#         merged_model[1].to_excel(f2)
# """

        # calculate distance between 0 and 1
        cluster_1 = merged_model[0]["Mean"]
        cluster_2 = merged_model[1]["Mean"]

        for k in range(len(cluster_1)):
            for j in range(len(cluster_2)):
                distances.append(distance.euclidean(cluster_1[k], cluster_2[j]))

        distances = np.array(distances).reshape(len(cluster_1), len(cluster_2))
        distance_df = pd.DataFrame(distances, columns=merged_model[1]["Cluster"],
                                   index=merged_model[0]["Cluster"])
        print("distance_df\n", distance_df)

        return distance_df

    return pd.DataFrame({})


def possible_cluster_merge(cluster_c1, cluster_c2, consolidation=False):
    model_num = 1
    if consolidation:
        model_num = 0
    n1 = merged_model[0].loc[cluster_c1]["No_of_pts"]
    n2 = merged_model[model_num].loc[cluster_c2]["No_of_pts"]
    n = n1 + n2
    c1 = np.array(merged_model[0].loc[cluster_c1]["Covariance"])
    c2 = np.array(merged_model[model_num].loc[cluster_c2]["Covariance"])
    m1 = np.array(merged_model[0].loc[cluster_c1]["Mean"])
    m2 = np.array(merged_model[model_num].loc[cluster_c2]["Mean"])
    m = (n1 * m1 + n2 * m2) / n

    merged_cov = (n1 * c1 + n2 * c2 + n1 * np.dot((m1 - m).transpose(), (m1 - m)) + n2 * np.dot((m2 - m).transpose(),
                                                                                                (m2 - m))) / n

    merged_entropy = 0.5 * (len(merged_cov) * log2(2 * pi * e) + log2(abs(np.linalg.det(merged_cov))))

    dist = distance.euclidean(m1, m2)

    delta_entropy = merged_entropy - max(merged_model[0].loc[cluster_c1]["Entropy"],
                                         merged_model[model_num].loc[cluster_c2]["Entropy"])
    entropy_perc = (delta_entropy / max(merged_model[0].loc[cluster_c1]["Entropy"],
                                        merged_model[model_num].loc[cluster_c2]["Entropy"])) * 100
    delta_2 = merged_entropy - min(merged_model[0].loc[cluster_c1]["Entropy"],
                                   merged_model[model_num].loc[cluster_c2]["Entropy"])
    entropy_2 = (delta_2 / min(merged_model[0].loc[cluster_c1]["Entropy"],
                               merged_model[model_num].loc[cluster_c2]["Entropy"])) * 100

    return merged_entropy, dist, delta_entropy, entropy_perc, delta_2, entropy_2, c1, c2, merged_cov


def create_sheet_(filename, option, index, write_df, styling=False):

    top = 5
    
    op = {0: "Merge_", 1: "MergeTop_", 2: "Consol_", 3: "ConsolTopBefore_", 4: "ConsolTopAfter_",
          5: "ConsBeforeRemove_"}
    wb = openpyxl.load_workbook(filename)
    sheet_name = op.get(option) + str(index)
    print("Sheet_name: ggg", sheet_name)
    ws = wb.create_sheet(sheet_name)
    wb.save(filename)

    wb = openpyxl.load_workbook(filename)
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        writer.book = wb
        writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
        write_df.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = wb[sheet_name]
        wb.active = ws

        if styling:

            alignment = Alignment(horizontal="center", vertical="center", wrapText=True, wrap_text=True)
            border_type = Side(border_style='thick', color='FF000000')
            border = Border(left=border_type,
                            right=border_type,
                            top=border_type,
                            bottom=border_type,
                            diagonal=border_type,
                            diagonal_direction=0,
                            outline=border_type,
                            vertical=border_type,
                            horizontal=border_type
                            )

            thick = Side(border_style="thick", color="000000")  # Border style, color
            border = Border(bottom=thick)  # Position of border
            row_range = list()
            start = 1
            for k in range(int(write_df.shape[0] / 5) + 1):
                a = "A" + str(start) + ":O" + str(start)
                start = start + top
                row_range.append(a)

            for cell_range in row_range:
                for row in ws[cell_range]:
                    for cell in row:
                        cell.border = border

            a = "A{0}:B{1}".format(write_df.shape[0], write_df.shape[0])
            b = "C{0}:K{1}".format(write_df.shape[0], write_df.shape[0])
            c = "L{0}:O{1}".format(write_df.shape[0], write_df.shape[0])
            cols = {a: 10, b: 15, c: 35}

            # for cell_range, size in cols.items():
            #     for row in ws[cell_range]:
            #         for cell in row:
            #             cell.alignment = alignment
            # cell.alignment.wrap_text = True
            # cell.alignment.horizontal = "center"
            # cell.alignment.vertical = "center"

            col1 = ['A', 'B']
            col2 = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
            col3 = ['K', 'L', 'M', 'O']

            for i in col1:
                ws.column_dimensions[i].width = 10
            for i in col2:
                ws.column_dimensions[i].width = 15
            for i in col3:
                ws.column_dimensions[i].width = 35

            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = alignment

            ws.freeze_panes = "A2"

    writer.save()
    wb.save(filename)

    return sheet_name


def build_comparison_matrix(distance_df, index, top=5):
    data = []
    ordered = np.argsort(distance_df.to_numpy())
    print("Ordered: \n", ordered)
    print("merged_0: ", merged_model[0])

    for r in range(ordered.shape[0]):
        dataset1_cluster = r
        # print("dataset1_cluster: \t", dataset1_cluster)
        for i in ordered[r][:top]:
            a = []
            # print(distance_df.iloc[r].name)
            # print(distance_df.columns[i])
            # print(r)
            # print( merged_model[0].iloc[r]["Entropy"])
            # print(merged_model[1].iloc[i]["Entropy"])
            a.extend([distance_df.iloc[r].name, distance_df.columns[i], merged_model[0].iloc[r]["Entropy"],
                      merged_model[1].iloc[i]["Entropy"]])
            a.extend([*possible_cluster_merge(dataset1_cluster, i, False)])
            data.append(a)

    df_entropy = pd.DataFrame(np.array(data, dtype=object), columns=["Dataset1", "Dataset2", "Entropy_Dataset1",
                                                                     "Entropy_Dataset2", "Merged_Entropy", "Distance",
                                                                     "Delta_Entropy", "DeltaEntropy%", "Delta_2",
                                                                     "Delta_2%",
                                                                     "Dataset1_covariance", "Dataset2_covariance",
                                                                     "Merged_covariance"])
    # merged_model[0].to_excel("merged_model_0.xlsx")
    # merged_model[1].to_excel("merged_model_1.xlsx")
# """
#     if index == 1:
#         if os.path.exists(os.getcwd() + '\\' + table_file):
#             os.remove(table_file)
#         w = xlsxwriter.Workbook(table_file)
#         w.add_worksheet()
#         w.close()
#
#     sheet = create_sheet_(table_file, 0, index, df_entropy, styling=True)
#     # view.style_excel(df_entropy, table_file, sheet, index=index, top=5)
# """
    return df_entropy


def merge(df_entropy, distance_df, df_ntop, index, consolidation=False):
    global cluster_instants
    # Create a dataframe containing the most recent
    # updates after merging

    sns.set(rc={'figure.figsize': (10, 6)})
    # merged_model[0].to_excel("merge_model0_old.xlsx")
# """
#     sheet = create_sheet_(merge_ntop, 1, index, df_ntop)
# """
    base = "DeltaEntropy%"
    count = 0
    model0 = merged_model[0]
    model1 = merged_model[1]
    datapoints[0]['data'] = pd.concat([datapoints[0]['data'], datapoints[1]['data']], ignore_index=True)
    data = datapoints[0]['data']
    label_list = list(set(data['id_label']))

    remove = []
    l = 0
    for k, rows in df_ntop.iterrows():
        c1 = rows['Dataset1']
        c2 = rows['Dataset2']
        if c1 in df_ntop['Dataset2'].values:
            row_index = df_ntop.index[df_ntop.Dataset2 == c1].values
            row_number = df_ntop.index.get_loc(df_ntop[df_ntop.Dataset2 == c1].iloc[-1].name)

            # check if 2 clusters are mutually close together.
            # e2 - [e2,a1,a2,a3,a4] a1 - [a1,a2,a3,e7,d3]
            # dont choose e2-a1 then.
            if df_entropy.loc[(df_entropy['Dataset1'] == c2) & (df_entropy['Dataset2'] == c1)].empty:
                remove.append(k)

            if k not in remove:
                if df_ntop.iloc[[row_number]][base].item() > rows[base]:
                    remove.append(row_index[0])
                else:
                    # At this point, you only have 1 element in row_index as
                    # we have removed duplicates in each column.
                    remove.append(k)

        l = l + 1

    df_ntop.drop(remove, axis=0, inplace=True)

    for k, rows in df_ntop.iterrows():
        c1 = rows["Dataset1"]
        c2 = rows["Dataset2"]

        # Save information
        m1_pts = model1[model1["Cluster"] == c2]["No_of_pts"]
        m1_mean = model1[model1["Cluster"] == c2]["Mean"]
        m0_pts = model0[model0["Cluster"] == c1]["No_of_pts"]
        m0_mean = model0[model0["Cluster"] == c1]["Mean"]

        # print("\n C1,C2: \t", c1,c2)
        # print("\n m0_m1: \t", m0_pts, m1_pts)
        mean = np.mean([m0_mean, m1_mean], axis=0)
        n_pts = m1_pts.item() + m0_pts.item()

        model0.loc[model0["Cluster"] == c1, ['No_of_pts']] = n_pts
        model0.loc[model0["Cluster"] == c1, ['Mean']] = mean
        model0[model0['Cluster'] == c1]['Covariance'].item()[:, :] = df_entropy.iloc[k]["Merged_covariance"]
        model0.loc[model0["Cluster"] == c1, ['Entropy']] = df_entropy.iloc[k]["Merged_Entropy"]
        model0.loc[model0["Cluster"] == c1, ['Cluster']] = "{0}{1}".format(char_list[index], str(count))
        d = "{0}{1}".format(char_list[index], str(count))
        count = count + 1
        notations[d] = str(c1) + "-" + str(c2)
        data.loc[(data['id_label'] == c1) | (data['id_label'] == c2), ['id_label']] = d

        # data.loc[(data['id_label'] == c2), ['label_2']] = data.loc[(data['id_label'] == c1), ['label_2']]

    n_cols = {23: 1, 46: 2, 92: 3, 184: 4}

    n_col = 1
    right = 0.7
    for m in n_cols.keys():
        if len(set(data['id_label'])) <= m:
            n_col = n_cols.get(m)
            if n_col > 2:
                right = 0.65
            break

    plt.clf()
    fig, ax = plt.subplots(figsize=(10, 6))
    g = sns.scatterplot(data=data, x="x", y="y", hue="id_label", s=15,
                        palette=view.generate_colors(len(set(data['id_label']))))
    plt.legend(bbox_to_anchor=(1.04, 1), borderaxespad=0, ncol=n_col)
    plt.setp(g.get_legend().get_texts(), fontsize='12')
    mean_np = np.array(model0["Mean"].tolist())
    plt.scatter(mean_np[:, 0], mean_np[:, 1], s=15, color="black", marker="+")
    s = "No. of clusters = " + str(len(set(data['id_label'])))
    plt.title(s)

    model1 = model1[~(model1["Cluster"].isin(df_ntop["Dataset2"]))]
    model0 = pd.concat([model0, model1], ignore_index=True)
    merged_model[0] = model0

    # Update cluster_instant dict for time tracing
    cluster_instants['merge_' + str(index)] = model0['Cluster']
    keys = natsort.natsorted(cluster_instants.keys())
    cluster_instants = OrderedDict((k, cluster_instants.get(k)) for k in keys)

    # model1.to_excel("merged_model1_new.xlsx")
    # merged_model[0].to_excel("merged_model_0.xlsx")

    for m in range(len(model0)):
        plt.text(model0.loc[m, "Mean"][0], model0.loc[m, "Mean"][1], model0.loc[m, "Cluster"],
                 horizontalalignment='center', verticalalignment='center', size=8, weight='bold',
                 color='black', backgroundcolor='white')

    plt.xlim(xmin, xmax)
    plt.ylim(ymin, ymax)
    merged_name = 'Merged_' + str(index) + ".png"
    plt.tight_layout(rect=[0, 0, 0.75, 1])

    plt.figtext(0.5, 0.01, "no. of points=" + str(model0['No_of_pts'].sum()), ha="center", fontsize=10,
                bbox={"facecolor": "orange", "alpha": 0.5, "pad": 5})

    plt.subplots_adjust(right=right)
    plt.savefig(merged_name, bbox_inches="tight")
    plt.close()
    # plt.show()

# """
#     if index == 1:
#         if os.path.exists(os.getcwd() + '\\' + merge_file):
#             os.remove(merge_file)
#         w = xlsxwriter.Workbook(merge_file)
#         w.add_worksheet()
#         w.close()
# """
# """
#     #   Write Images to excel
#     wb = openpyxl.load_workbook(merge_file)
#     print("Merge file 123: \n", merge_file)
#     sheet_name = "MergeRound" + str(index)
#     print("Sheet_name: ddd", sheet_name)
#     ws = wb.create_sheet(sheet_name)
#     images = {"A2": datapoints[0]['image'], "J2": datapoints[1]['image'], "S2": merged_name}
#
#     max = 2
#     for k, v in notations.items():
#         ws.cell(row=max, column=28).value = str(k)
#         ws.cell(row=max, column=29).value = str(v)
#         max = max + 1
#
#     wb.save(merge_file)
#
#     for k, v in images.items():
#         img = openpyxl.drawing.image.Image(v)
#         img.anchor = k
#         img.width = 571.2
#         img.height = 555.84
#         ws.add_image(img)
#
#     wb.save(merge_file)
#
#     for ws in wb.worksheets:
#         ws.sheet_view.zoomScale = 83.2
#         if ws.title == "Sheet1":
#             ws.sheet_view.zoomScale = 95
#
#     wb.save(merge_file)
# """

    datapoints[0]['image'] = merged_name


def consolidate():
    # calculate distance
    consol_dist = []

    merged_model[0] = merged_model[0].sort_values(by="Cluster", key=natsort.natsort_keygen(), ignore_index=True)
    cluster = merged_model[0]['Mean']

    for k in range(len(cluster)):
        for j in range(len(cluster)):
            consol_dist.append(distance.euclidean(cluster[k], cluster[j]))

    consol_dist = np.array(consol_dist).reshape(len(cluster), len(cluster))
    consol_df = pd.DataFrame(consol_dist, columns=merged_model[0]['Cluster'], index=merged_model[0]['Cluster'])
    # print("Consolidated_df: \n", consol_df)

    data = []
    ordered = np.argsort(consol_df.to_numpy())
    # print("Ordered: \n", ordered)

    for r in range(ordered.shape[0]):
        dataset1_cluster = r
        # print("dataset1_cluster: \t", dataset1_cluster)
        for i in ordered[r][:6]:
            a = []
            a.extend([consol_df.iloc[r].name, consol_df.columns[i], merged_model[0].iloc[r]["Entropy"],
                      merged_model[0].iloc[i]["Entropy"]])
            a.extend([*possible_cluster_merge(dataset1_cluster, i, True)])
            data.append(a)

    df_entropy_cons = pd.DataFrame(np.array(data, dtype=object), columns=["Dataset1", "Dataset2", "Entropy_Dataset1",
                                                                          "Entropy_Dataset2", "Merged_Entropy",
                                                                          "Distance",
                                                                          "Delta_Entropy", "DeltaEntropy%", "Delta_2",
                                                                          "Delta_2%",
                                                                          "Dataset1_covariance", "Dataset2_covariance",
                                                                          "Merged_covariance"])

    return df_entropy_cons


def consolidate_plot(index):
    global cluster_instants

    df_entropy_cons = consolidate()
    model0 = merged_model[0]
    count = 0
    # base = "DeltaEntropy%"
    base_2 ="DeltaEntropy%"
    base = "Delta_2%"
    data = datapoints[0]['data']
    ntop_file = cons_ntop

# """
#     if index == 1:
#         if os.path.exists(os.getcwd() + '\\' + ntop_file):
#             os.remove(ntop_file)
#         w = xlsxwriter.Workbook(ntop_file)
#         w.add_worksheet()
#         w.close()
# """
    if not df_entropy_cons.empty:
        try:
        # """
        #     if index == 1:
        #         if os.path.exists(os.getcwd() + '\\' + cons_entropy):
        #             os.remove(cons_entropy)
        #         w = xlsxwriter.Workbook(cons_entropy)
        #         worksheet = w.add_worksheet()
        #         w.close()
        # """

            df_entropy_cons = df_entropy_cons[(df_entropy_cons['Dataset1'] != df_entropy_cons['Dataset2'])]
            df_entropy_cons.reset_index(drop=True, inplace=True)
        # """
        #     sheet = create_sheet_(cons_entropy, 2, index, df_entropy_cons, styling=True)
        # """
            # view.style_excel(df_entropy_cons, "Consolidate_entropy_1.xlsx",sheet=sheet,index=index, top=5)

            # print('df_entropy_cons: \n', df_entropy_cons)
            # & (df_entropy_cons[base] > -5.2)
            df_ntop_cons = df_entropy_cons[(df_entropy_cons[base_2] < -2) & (df_entropy_cons[base_2] > -75) &
                                           (df_entropy_cons[base] < -6) & (df_entropy_cons[base_2] > 6)&  (
                    df_entropy_cons['Dataset1'] != df_entropy_cons['Dataset2'])][
                ["Dataset1", "Dataset2", "Distance", base, base_2]]
            df_ntop_cons = df_ntop_cons.sort_values(by=["Distance"], ascending=True).head(30)
            print("Consolidation: \n", df_ntop_cons)

            # Remove self-merges
            if not df_ntop_cons.empty:
                df_ntop_cons = df_ntop_cons[df_ntop_cons['Dataset1'] != df_ntop_cons['Dataset2']]

            if not df_ntop_cons.empty:
                # Drop reverse duplicates
                df_ntop_cons['sorted_row'] = df_ntop_cons.apply(
                    lambda row: ''.join(sorted([row['Dataset1'], row['Dataset2']])), axis=1)
                df_ntop_cons.drop_duplicates(subset=['sorted_row'], inplace=True)
                df_ntop_cons.drop(columns=['sorted_row'], inplace=True)

        # """
        #     sheet = create_sheet_(cons_ntop, 5, index, df_ntop_cons)
        # """

            # Remove duplicates across d1 and d2 separately
            # Keep = "first" because the entropy% is ordered.
            d1 = df_ntop_cons[df_ntop_cons.duplicated("Dataset1", keep="first")]
            df_ntop_cons = df_ntop_cons[~df_ntop_cons.isin(d1)].dropna(how="all")
            d2 = df_ntop_cons[df_ntop_cons.duplicated("Dataset2", keep="first")]
            df_ntop_cons = df_ntop_cons[~df_ntop_cons.isin(d2)].dropna(how="all")

            # Remove instance of D0_1 - D1_5 when D1_5 - D0_7 exists
            # Criss-cross single instance removal

            # Iterate through rows, merge and eliminate instances in the further rows
            # ex : C4 - D0_1 and D0_1 - E4
            remove = []
            l = 0
            for k, rows in df_ntop_cons.iterrows():
                c1 = rows['Dataset1']
                c2 = rows['Dataset2']
                if c1 in df_ntop_cons['Dataset2'].values:
                    row_index = df_ntop_cons.index[df_ntop_cons.Dataset2 == c1].values
                    row_number = df_ntop_cons.index.get_loc(df_ntop_cons[df_ntop_cons.Dataset2 == c1].iloc[-1].name)

                    # check if 2 clusters are mutually close together.
                    # e2 - [e2,a1,a2,a3,a4] a1 - [a1,a2,a3,e7,d3]
                    # dont choose e2-a1 then.
                    if df_entropy_cons.loc[
                        (df_entropy_cons['Dataset1'] == c2) & (df_entropy_cons['Dataset2'] == c1)].empty:
                        remove.append(k)

                    if k not in remove:
                        if df_ntop_cons.iloc[[row_number]][base].item() > rows[base]:
                            remove.append(row_index[0])
                        else:
                            # At this point, you only have 1 element in row_index as
                            # we have removed duplicates in each column.
                            remove.append(k)

                l = l + 1
                # if df_ntop_cons.loc[[row_number[0]]]['DeltaEntropy%'] > rows['DeltaEntropy%']:
                #     remove.append(k)
        # """
        #     sheet = create_sheet_(cons_ntop, 3, index, df_ntop_cons)
        # """

            print("remove:", remove)
            # drop the rows
            df_ntop_cons.drop(remove, axis=0, inplace=True)
        # """
        #     sheet = create_sheet_(cons_ntop, 4, index, df_ntop_cons)
        # """
            # df_ntop_cons.to_excel("df_ntop_cons_test.xlsx")

            # datapoints[0]['data'] contains all data points we need for consolidation.
            # Plot and update mean

            consol_model[0]['df'] = merged_model[0].copy()
            consol_model[0]['image'] = datapoints[0]['image']
            consol_model[0]['data'] = datapoints[0]['data']

            consol_model[1]['df'] = merged_model[0].copy()
            consol_model[1]['df']['name'] = np.nan
            consol_model[1]['data'] = datapoints[0]['data']

            count = 0
            for k, rows in df_ntop_cons.iterrows():
                c1 = rows["Dataset1"]
                c2 = rows["Dataset2"]

                # Save information
                m1_pts = model0[model0["Cluster"] == c2]["No_of_pts"]
                m1_mean = model0[model0["Cluster"] == c2]["Mean"]
                m0_pts = model0[model0["Cluster"] == c1]["No_of_pts"]
                m0_mean = model0[model0["Cluster"] == c1]["Mean"]

                # print("\n consol, c1,c2: \t", c1,c2)
                mean = np.mean([m0_mean, m1_mean], axis=0)
                n_pts = int(m1_pts.item() + m0_pts.item())

                consol_model[1]['df'].loc[consol_model[1]['df']['Cluster'] == c1, ['No_of_pts']] = n_pts
                consol_model[1]['df'].loc[consol_model[1]['df']["Cluster"] == c1, ['Mean']] = mean

                np.array(consol_model[1]['df'][consol_model[1]['df']['Cluster'] == c1]['Covariance'].item())[:, :] = \
                    df_entropy_cons.iloc[k]["Merged_covariance"]
                consol_model[1]['df'].loc[consol_model[1]['df']["Cluster"] == c1, ['Entropy']] = \
                    df_entropy_cons.iloc[k][
                        "Merged_Entropy"]

                consol_model[1]['df'].loc[consol_model[1]['df']["Cluster"] == c1, ['name']] = '{0}{1}'.format(
                    cons_list[index], str(count))

                consol_model[1]['df'].loc[consol_model[1]['df']["Cluster"] == c1, ['Cluster']] = "{0}-{1}".format(c1,
                                                                                                                  c2)

                consol_model[1]['df'] = consol_model[1]['df'][
                    ~(consol_model[1]['df']["Cluster"].isin(df_ntop_cons["Dataset2"]))]

                d = "{0}-{1}".format(c1, c2)

                consol_model[1]['data'].loc[
                    (consol_model[1]['data']['id_label'] == c1) | (consol_model[1]['data']['id_label'] == c2), [
                        'id_label']] = d

                consol_model[1]['df'].reset_index(drop=True, inplace=True)

                count = count + 1

            n_cols = {23: 1, 46: 2, 92: 3, 184: 4}
            # print("\n data:\t", consol_model[1])
            n_col = 1
            right = 0.7
            for m in n_cols.keys():
                if len(set(consol_model[1]['data']['id_label'])) <= m:
                    n_col = n_cols.get(m)
                    if n_col > 2:
                        right = 0.65
                    break

            plt.clf()
            fig, ax = plt.subplots(figsize=(10, 6))
            g = sns.scatterplot(data=data, x="x", y="y", hue="id_label", s=15,
                                palette=view.generate_colors(len(set(consol_model[1]['data']['id_label']))))
            plt.legend(bbox_to_anchor=(1.04, 1), borderaxespad=0, ncol=n_col)
            plt.setp(g.get_legend().get_texts(), fontsize='12')

            mean_np = np.array(consol_model[1]['df']['Mean'].tolist())
            plt.scatter(mean_np[:, 0], mean_np[:, 1], s=15, color="black", marker="+")
            s = "No. of clusters = " + str(len(set(consol_model[1]['data']['id_label'])))
            plt.title(s)

            for m in range(len(consol_model[1]['df'])):
                plt.text(consol_model[1]['df'].loc[m, "Mean"][0], consol_model[1]['df'].loc[m, "Mean"][1],
                         consol_model[1]['df'].loc[m, "Cluster"],
                         horizontalalignment='center', verticalalignment='center', size=8, weight='bold',
                         color='black', backgroundcolor='white')

            plt.xlim(xmin, xmax)
            plt.ylim(ymin, ymax)

            merged_name = 'Cons_' + str(index) + ".png"

            plt.tight_layout(rect=[0, 0, 0.85, 1])
            plt.figtext(0.5, 0.01, "no. of points=" + str(consol_model[1]['df']['No_of_pts'].sum()), ha="center",
                        fontsize=10,
                        bbox={"facecolor": "orange", "alpha": 0.5, "pad": 5})
            plt.subplots_adjust(right=right)
            plt.savefig(merged_name, bbox_inches="tight")
            plt.close()

        # """
        #     #   Write Images to excel
        #     wb = openpyxl.load_workbook(merge_file)
        #     sheet_name = "Consolidate_" + str(index)
        #     print("Sheet_name:yyy", sheet_name)
        #     ws = wb.create_sheet(sheet_name)
        #     images = {"A2": consol_model[0]['image'], "O2": merged_name}
        #
        #     max = 2
        #     for k, v in notations.items():
        #         ws.cell(row=max, column=28).value = str(k)
        #         ws.cell(row=max, column=29).value = str(v)
        #         max = max + 1
        #     wb.save(merge_file)
        #
        #     img_w = 873.6 if n_col > 1 else 571.2
        #
        #     for k, v in images.items():
        #         img = openpyxl.drawing.image.Image(v)
        #         img.anchor = k
        #         img.width = img_w
        #         img.height = 576
        #         ws.add_image(img)
        #
        #     wb.save(merge_file)
        #
        #     for ws in wb.worksheets:
        #         ws.sheet_view.zoomScale = 78
        #         if ws.title == "Sheet1":
        #             ws.sheet_view.zoomScale = 95
        #
        #     wb.save(merge_file)
        # """

            # shorten the names
            slice_df = consol_model[1]['df'][consol_model[1]['df']['name'].notnull()]
            slice_df.reset_index(drop=True, inplace=True)
            consol_model[1]['data'].reset_index(drop=True, inplace=True)

            for k, rows in slice_df.iterrows():
                c1 = rows['Cluster']
                name = rows['name']
                consol_model[1]['data'].loc[consol_model[1]['data']['id_label'] == c1, ['id_label']] = rows['name']
                # consol_model[1]['data'][['id_label']] = consol_model[1]['data'][['id_label']].where(
                #     consol_model[1]['data']['id_label'] == c1, s)

            datapoints[0]['data'] = consol_model[1]['data']
            data = datapoints[0]['data']

            # consol_model[1]['data'].loc[consol_model[1]['data']['id_label'].isin(slice_df['Cluster']), ['id_label']] = \
            # slice_df[['name']]

            print("before: \t", consol_model[1]['df'][consol_model[1]['df']['name'].notnull()])
            consol_model[1]['df']['Cluster'] = np.where(consol_model[1]['df']['name'].notnull(),
                                                        consol_model[1]['df']['name'], consol_model[1]['df']['Cluster'])
            print("after: \t", consol_model[1]['df'][consol_model[1]['df']['name'].notnull()])

            cons_dict = dict(zip(slice_df['name'], slice_df['Cluster']))
            notations.update(cons_dict)

            n_cols = {23: 1, 46: 2, 92: 3, 184: 4}
            # print("\n data:\t", consol_model[1])
            n_col = 1
            right = 0.7
            for m in n_cols.keys():
                if len(set(consol_model[1]['data']['id_label'])) <= m:
                    n_col = n_cols.get(m)
                    if n_col > 2:
                        right = 0.65
                    break

            plt.clf()
            fig, ax = plt.subplots(figsize=(10, 6))

            g = sns.scatterplot(data=data, x="x", y="y", hue="id_label", s=15,
                                palette=view.generate_colors(len(set(consol_model[1]['data']['id_label']))))
            plt.legend(bbox_to_anchor=(1.04, 1), borderaxespad=0, ncol=n_col)
            plt.setp(g.get_legend().get_texts(), fontsize='12')

            mean_np = np.array(consol_model[1]['df']['Mean'].tolist())
            plt.scatter(mean_np[:, 0], mean_np[:, 1], s=15, color="black", marker="+")
            s = "No. of clusters = " + str(len(set(consol_model[1]['data']['id_label'])))
            plt.title(s)

            for m in range(len(consol_model[1]['df'])):
                plt.text(consol_model[1]['df'].loc[m, "Mean"][0], consol_model[1]['df'].loc[m, "Mean"][1],
                         consol_model[1]['df'].loc[m, "Cluster"],
                         horizontalalignment='center', verticalalignment='center', size=8, weight='bold',
                         color='black', backgroundcolor='white')

            plt.xlim(xmin, xmax)
            plt.ylim(ymin, ymax)

            merged_name = 'Cons_new_' + str(index) + ".png"

            plt.tight_layout(rect=[0, 0, 0.75, 1])
            plt.figtext(0.5, 0.01, "no. of points=" + str(consol_model[1]['df']['No_of_pts'].sum()), ha="center",
                        fontsize=10,
                        bbox={"facecolor": "orange", "alpha": 0.5, "pad": 5})
            plt.subplots_adjust(right=right)
            plt.savefig(merged_name, bbox_inches="tight")
            plt.close()

            consol_model[1]['image'] = merged_name
            consol_model[0]['image'] = merged_name
            merged_model[0] = consol_model[1]['df']
            datapoints[0]['image'] = consol_model[1]['image']
            datapoints[0]['data'] = consol_model[1]['data']

            # Update cluster_instants for time tracing
            # Get last merge key
            ind = list(cluster_instants.keys())[-1]
            cluster_instants[ind] = merged_model[0]['Cluster']
            keys = natsort.natsorted(cluster_instants.keys())
            cluster_instants = OrderedDict((k, cluster_instants.get(k)) for k in keys)
            #
            # if index == 3:
            #     consol_model[1]['df'].to_excel("DISHA_AFTER_1.xlsx")
            #     consol_model[1]['data'].to_excel("DISHA_AFTER_2.xlsx")

        except ValueError as e:
            print("Error!", e)
            traceback.print_exc()
            sys.exit(1)

        except FileNotFoundError as e:
            print("File not found!")
            traceback.print_exc()
            sys.exit(1)

        except KeyError as e:
            print("KEYERROR")
            # print(e.message)
            print(e)
            traceback.print_exc()


def consolidate_model(index, df_ntop_cons, df_entropy_cons):
    global cluster_instants
    model0 = merged_model[0]
    count = 0
    # base = "DeltaEntropy%"
    base = "Delta_2%"
    data = datapoints[0]['data']
    ntop_file = cons_ntop

    consol_model[0]['df'] = merged_model[0].copy()
    consol_model[0]['image'] = datapoints[0]['image']
    consol_model[0]['data'] = datapoints[0]['data']

    consol_model[1]['df'] = merged_model[0].copy()
    consol_model[1]['df']['name'] = np.nan
    consol_model[1]['data'] = datapoints[0]['data']

    consol_model[1]['df'].to_excel("before_cons.xlsx")

    count = 0
    for k, rows in df_ntop_cons.iterrows():

        c1 = rows["Dataset1"]
        c2 = rows["Dataset2"]

        # Save information
        m1_pts = model0[model0["Cluster"] == c2]["No_of_pts"]
        m1_mean = model0[model0["Cluster"] == c2]["Mean"]
        m0_pts = model0[model0["Cluster"] == c1]["No_of_pts"]
        m0_mean = model0[model0["Cluster"] == c1]["Mean"]

        # print("\n consol, c1,c2: \t", c1,c2)
        mean = np.mean([m0_mean, m1_mean], axis=0)
        if not m1_pts.empty:
            print("cluster c2 ", c2)
        if not m0_pts.empty:
            print("cluster c1 ", c1)
        print('Check:', m1_pts)
        print("Check2:", m0_pts)
        n_pts = int(m1_pts.item() + m0_pts.item())

        consol_model[1]['df'].loc[consol_model[1]['df']['Cluster'] == c1, ['No_of_pts']] = n_pts
        consol_model[1]['df'].loc[consol_model[1]['df']["Cluster"] == c1, ['Mean']] = mean

        np.array(consol_model[1]['df'][consol_model[1]['df']['Cluster'] == c1]['Covariance'].item())[:, :] = \
            df_entropy_cons.iloc[k]["Merged_covariance"]
        consol_model[1]['df'].loc[consol_model[1]['df']["Cluster"] == c1, ['Entropy']] = df_entropy_cons.iloc[k][
            "Merged_Entropy"]

        consol_model[1]['df'].loc[consol_model[1]['df']["Cluster"] == c1, ['name']] = '{0}{1}'.format(cons_list[index],
                                                                                                      str(count))

        consol_model[1]['df'].loc[consol_model[1]['df']["Cluster"] == c1, ['Cluster']] = "{0}-{1}".format(c1, c2)

        consol_model[1]['df'] = consol_model[1]['df'][
            ~(consol_model[1]['df']["Cluster"].isin(df_ntop_cons["Dataset2"]))]

        d = "{0}-{1}".format(c1, c2)

        consol_model[1]['data'].loc[
            (consol_model[1]['data']['id_label'] == c1) | (consol_model[1]['data']['id_label'] == c2), [
                'id_label']] = d

        # consol_model[0]['df'].to_excel("consolidate_model_before.xlsx")
        # consol_model[1]['df'].to_excel("consolidate_model_df.xlsx")
        consol_model[1]['df'].reset_index(drop=True, inplace=True)
        consol_model[1]['df'].to_excel("consolidate_model_ar.xlsx")

        count = count + 1

        if index == 3:
            consol_model[1]['df'].to_excel("DISHA_BEFORE_1.xlsx")
            consol_model[1]['data'].to_excel("DISHA_BEFORE_2.xlsx")

    n_cols = {23: 1, 46: 2, 92: 3, 184: 4}
    # print("\n data:\t", consol_model[1])
    n_col = 1
    right = 0.7
    for m in n_cols.keys():
        if len(set(consol_model[1]['data']['id_label'])) <= m:
            n_col = n_cols.get(m)
            if n_col > 2:
                right = 0.65
            break

    plt.clf()
    fig, ax = plt.subplots(figsize=(10, 6))
    g = sns.scatterplot(data=data, x="x", y="y", hue="id_label", s=15,
                        palette=view.generate_colors(len(set(consol_model[1]['data']['id_label']))))
    plt.legend(bbox_to_anchor=(1.04, 1), borderaxespad=0, ncol=n_col)
    plt.setp(g.get_legend().get_texts(), fontsize='12')

    mean_np = np.array(consol_model[1]['df']['Mean'].tolist())
    plt.scatter(mean_np[:, 0], mean_np[:, 1], s=15, color="black", marker="+")
    s = "No. of clusters = " + str(len(set(consol_model[1]['data']['id_label'])))
    plt.title(s)

    for m in range(len(consol_model[1]['df'])):
        plt.text(consol_model[1]['df'].loc[m, "Mean"][0], consol_model[1]['df'].loc[m, "Mean"][1],
                 consol_model[1]['df'].loc[m, "Cluster"],
                 horizontalalignment='center', verticalalignment='center', size=8, weight='bold',
                 color='black', backgroundcolor='white')

    plt.xlim(xmin, xmax)
    plt.ylim(ymin, ymax)

    merged_name = 'Cons_' + str(index) + ".png"

    plt.tight_layout(rect=[0, 0, 0.75, 1])
    plt.figtext(0.5, 0.01, "no. of points=" + str(consol_model[1]['df']['No_of_pts'].sum()), ha="center", fontsize=10,
                bbox={"facecolor": "orange", "alpha": 0.5, "pad": 5})
    plt.subplots_adjust(right=right)
    plt.savefig(merged_name, bbox_inches="tight")
    plt.close()

# """
#     #   Write Images to excel
#     wb = openpyxl.load_workbook(merge_file)
#     sheet_name = "Consolidate_" + str(index)
#     print("Sheet_name:yyy", sheet_name)
#     ws = wb.create_sheet(sheet_name)
#     images = {"A2": consol_model[0]['image'], "O2": merged_name}
#
#     max = 2
#     for k, v in notations.items():
#         ws.cell(row=max, column=28).value = str(k)
#         ws.cell(row=max, column=29).value = str(v)
#         max = max + 1
#     wb.save(merge_file)
#
#     # img_w = 873.6 if n_col > 1 else 571.2
#     img_w = 873.6
#     for k, v in images.items():
#         img = openpyxl.drawing.image.Image(v)
#         img.anchor = k
#         img.width = img_w
#         img.height = 555.84
#         ws.add_image(img)
#
#     wb.save(merge_file)
#
#     for ws in wb.worksheets:
#         ws.sheet_view.zoomScale = 78
#         if ws.title == "Sheet1":
#             ws.sheet_view.zoomScale = 95
#
#     wb.save(merge_file)
# """

    # shorten the names
    slice_df = consol_model[1]['df'][consol_model[1]['df']['name'].notnull()]
    slice_df.reset_index(drop=True, inplace=True)
    consol_model[1]['data'].reset_index(drop=True, inplace=True)

    for k, rows in slice_df.iterrows():
        c1 = rows['Cluster']
        name = rows['name']
        consol_model[1]['data'].loc[consol_model[1]['data']['id_label'] == c1, ['id_label']] = rows['name']
        # consol_model[1]['data'][['id_label']] = consol_model[1]['data'][['id_label']].where(
        #     consol_model[1]['data']['id_label'] == c1, s)

    datapoints[0]['data'] = consol_model[1]['data']
    data = datapoints[0]['data']

    # consol_model[1]['data'].loc[consol_model[1]['data']['id_label'].isin(slice_df['Cluster']), ['id_label']] = \
    # slice_df[['name']]

    print("before: \t", consol_model[1]['df'][consol_model[1]['df']['name'].notnull()])
    consol_model[1]['df']['Cluster'] = np.where(consol_model[1]['df']['name'].notnull(), consol_model[1]['df']['name'],
                                                consol_model[1]['df']['Cluster'])
    print("after: \t", consol_model[1]['df'][consol_model[1]['df']['name'].notnull()])

    cons_dict = dict(zip(slice_df['name'], slice_df['Cluster']))
    notations.update(cons_dict)

    n_cols = {23: 1, 46: 2, 92: 3, 184: 4}
    # print("\n data:\t", consol_model[1])
    n_col = 1
    right = 0.7
    for m in n_cols.keys():
        if len(set(consol_model[1]['data']['id_label'])) <= m:
            n_col = n_cols.get(m)
            if n_col > 2:
                right = 0.65
            break

    plt.clf()
    fig, ax = plt.subplots(figsize=(10, 6))

    g = sns.scatterplot(data=data, x="x", y="y", hue="id_label", s=15,
                        palette=view.generate_colors(len(set(consol_model[1]['data']['id_label']))))
    plt.legend(bbox_to_anchor=(1.04, 1), borderaxespad=0, ncol=n_col)
    plt.setp(g.get_legend().get_texts(), fontsize='12')

    mean_np = np.array(consol_model[1]['df']['Mean'].tolist())
    plt.scatter(mean_np[:, 0], mean_np[:, 1], s=15, color="black", marker="+")
    s = "No. of clusters = " + str(len(set(consol_model[1]['data']['id_label'])))
    plt.title(s)

    for m in range(len(consol_model[1]['df'])):
        plt.text(consol_model[1]['df'].loc[m, "Mean"][0], consol_model[1]['df'].loc[m, "Mean"][1],
                 consol_model[1]['df'].loc[m, "Cluster"],
                 horizontalalignment='center', verticalalignment='center', size=8, weight='bold',
                 color='black', backgroundcolor='white')

    plt.xlim(xmin, xmax)
    plt.ylim(ymin, ymax)

    merged_name = 'Cons_new_' + str(index) + ".png"

    plt.tight_layout(rect=[0, 0, 0.75, 1])
    plt.figtext(0.5, 0.01, "no. of points=" + str(consol_model[1]['df']['No_of_pts'].sum()), ha="center", fontsize=10,
                bbox={"facecolor": "orange", "alpha": 0.5, "pad": 5})
    plt.subplots_adjust(right=right)
    plt.savefig(merged_name, bbox_inches="tight")
    plt.close()

    consol_model[1]['image'] = merged_name
    consol_model[0]['image'] = merged_name
    merged_model[0] = consol_model[1]['df']
    datapoints[0]['image'] = consol_model[1]['image']
    datapoints[0]['data'] = consol_model[1]['data']

    # Update cluster_instants for time tracing
    # Get last merge key
    ind = list(cluster_instants.keys())[-1]
    cluster_instants[ind] = merged_model[0]['Cluster']
    keys = natsort.natsorted(cluster_instants.keys())
    cluster_instants = OrderedDict((k, cluster_instants.get(k)) for k in keys)


def check_consolidation(index, df_entropy_cons):
    # df_entropy_cons = consolidate()
    model0 = merged_model[0]
    count = 0
    # base = "DeltaEntropy%"
    base_2 = "DeltaEntropy%"
    base = "Delta_2%"
    data = datapoints[0]['data']
    ntop_file = cons_ntop
# """
#     if index == 1:
#         if os.path.exists(os.getcwd() + '\\' + ntop_file):
#             os.remove(ntop_file)
#         w = xlsxwriter.Workbook(ntop_file)
#         w.add_worksheet()
#         w.close()
# """

    if not df_entropy_cons.empty:
        try:
            # if index == 1:
        #     """
        #         if os.path.exists(os.getcwd() + '\\' + cons_entropy):
        #             os.remove(cons_entropy)
        #         w = xlsxwriter.Workbook(cons_entropy)
        #         worksheet = w.add_worksheet()
        #         w.close()
        #     """
        # """
        #     sheet = create_sheet_(cons_entropy, 2, index, df_entropy_cons, styling=True)
        # """
            # view.style_excel(df_entropy_cons, "Consolidate_entropy_1.xlsx",sheet=sheet,index=index, top=5)

            # print('df_entropy_cons: \n', df_entropy_cons)
            # & (df_entropy_cons[base] > -5.2)
            df_ntop_cons = df_entropy_cons[(df_entropy_cons[base_2] < -2) & (df_entropy_cons[base_2] > -75) &
                                           (df_entropy_cons[base] < -6) & (df_entropy_cons[base_2] > 6)  & (
                    df_entropy_cons['Dataset1'] != df_entropy_cons['Dataset2'])][
                ["Dataset1", "Dataset2", base, base_2]]
            df_ntop_cons = df_ntop_cons.sort_values(by=[base], ascending=True).head(30)
            print("Consolidation: \n", df_ntop_cons)

            # Remove self-merges
            if not df_ntop_cons.empty:
                df_ntop_cons = df_ntop_cons[df_ntop_cons['Dataset1'] != df_ntop_cons['Dataset2']]

            if not df_ntop_cons.empty:
                # Drop reverse duplicates
                df_ntop_cons['sorted_row'] = df_ntop_cons.apply(
                    lambda row: ''.join(sorted([row['Dataset1'], row['Dataset2']])), axis=1)
                df_ntop_cons.drop_duplicates(subset=['sorted_row'], inplace=True)
                df_ntop_cons.drop(columns=['sorted_row'], inplace=True)
        # """
        #     sheet = create_sheet_(cons_ntop, 5, index, df_ntop_cons)
        # """
            # Remove duplicates across d1 and d2 separately
            # Keep = "first" because the entropy% is ordered.
            d1 = df_ntop_cons[df_ntop_cons.duplicated("Dataset1", keep="first")]
            df_ntop_cons = df_ntop_cons[~df_ntop_cons.isin(d1)].dropna(how="all")
            d2 = df_ntop_cons[df_ntop_cons.duplicated("Dataset2", keep="first")]
            df_ntop_cons = df_ntop_cons[~df_ntop_cons.isin(d2)].dropna(how="all")

            if df_ntop_cons.empty:
                return pd.DataFrame({})

            # Remove instance of D0_1 - D1_5 when D1_5 - D0_7 exists
            # Criss-cross single instance removal

            # Iterate through rows, merge and eliminate instances in the further rows
            # ex : C4 - D0_1 and D0_1 - E4
            remove = []
            l = 0
            for k, rows in df_ntop_cons.iterrows():
                c1 = rows['Dataset1']
                c2 = rows['Dataset2']
                if c1 in df_ntop_cons['Dataset2'].values:
                    row_index = df_ntop_cons.index[df_ntop_cons.Dataset2 == c1].values
                    row_number = df_ntop_cons.index.get_loc(df_ntop_cons[df_ntop_cons.Dataset2 == c1].iloc[-1].name)

                    # check if 2 clusters are mutually close together.
                    # e2 - [e2,a1,a2,a3,a4] a1 - [a1,a2,a3,e7,d3]
                    # dont choose e2-a1 then.
                    if df_entropy_cons.loc[
                        (df_entropy_cons['Dataset1'] == c2) & (df_entropy_cons['Dataset2'] == c1)].empty:
                        remove.append(k)

                    if k not in remove:
                        if df_ntop_cons.iloc[[row_number]][base].item() > rows[base]:
                            remove.append(row_index[0])
                        else:
                            # At this point, you only have 1 element in row_index as
                            # we have removed duplicates in each column.
                            remove.append(k)

                l = l + 1
                # if df_ntop_cons.loc[[row_number[0]]]['DeltaEntropy%'] > rows['DeltaEntropy%']:
                #     remove.append(k)
        # """
        #     sheet = create_sheet_(cons_ntop, 3, index, df_ntop_cons)
        # """
            print("remove:", remove)
            # drop the rows
            df_ntop_cons.drop(remove, axis=0, inplace=True)
        # """
        #     sheet = create_sheet_(cons_ntop, 4, index, df_ntop_cons)
        # """

            # consolidate_model(index, df_ntop_cons, df_entropy_cons)
            return df_ntop_cons

        except ValueError as e:
            print("Error!", e)
            traceback.print_exc()
            sys.exit(1)

        except FileNotFoundError as e:
            traceback.print_exc()
            sys.exit(1)

    else:
        return pd.DataFrame({})


def build_heatmap_matrix(alt_data):
    file = prefix + "_" + str(chunk_size) + "_compare_2.xlsx"
    result_n_label = alt_data['t_label_0'] #labels for k = variable number
    # n_15_label = alt_data['t_label_1'] #label for k= 50
    n_15_label = datapoints[0]['data']['truth']
    result_df_label = datapoints[0]['data']['labels']

    if os.path.exists(os.getcwd() + '\\' + file):
        os.remove(file)
    w = xlsxwriter.Workbook(file)
    worksheet = w.add_worksheet()
    w.close()

    s1 = "rand index at once whole with variable n = {0}  gmm vs Alg result :{1} ".format(
        str(len(set(alt_data['t_label_0']))), str(
            rand_score(result_n_label, result_df_label)))

    s2 = "rand index at once whole with fixed    n = {0}  gmm vs Alg result : {1} ".format(
        str(len(set(alt_data['t_label_1']))), str(
            rand_score(n_15_label, result_df_label)))

    wb = openpyxl.load_workbook(file)
    sheet_name = "Sheet1"
    ws = wb.create_sheet(sheet_name)

    image_dict = {'A3.xlsx': 'A3_original.png'}

    # full_data_gmm_0.png : whole <n> gmm
    # full_data_gmm_1.png : whole n=15 gmm
    images = {"A3": "Original.png" if dataset != 'A3.xlsx' else image_dict['A3.xlsx'], "J3": "full_data_gmm_0.png",
              "S3": consol_model[0]['image'],
              "AB3": "full_data_gmm_1.png"}

    ws.cell(row=33, column=12).value = s1
    ws.cell(row=34, column=12).value = s2
    ws.cell(row=1, column=1).value = "Website Graph"
    ws.cell(row=1, column=10).value = "GMM, At once variable n=" + str(
        len(set(alt_data['t_label_0']))) + "chunk_size=" + str(chunk_size)
    ws.cell(row=1, column=19).value = "Algorithm graph, chunk size=" + str(chunk_size)
    ws.cell(row=1, column=28).value = "GMM, At once fixed n=" + str(
        len(set(alt_data['t_label_1']))) + "chunk_size=" + str(chunk_size)

    wb.save(file)

    # img_w = 873.6 if n_col > 1 else 571.2
    # img_w = 873.6
    for k, v in images.items():
        img = openpyxl.drawing.image.Image(v)
        img.anchor = k
        img.width = 571.2
        img.height = 555.84
        ws.add_image(img)

    wb.save(file)

    for ws in wb.worksheets:
        ws.sheet_view.zoomScale = 78
        if ws.title == "Sheet1":
            ws.sheet_view.zoomScale = 95

    wb.save(file)

    return


def time_trace(instant: string):
    global cluster_instants
    tt = {}
    clusters = cluster_instants[instant]
    print("Length of clusters:", len(clusters))

    def find_parent(c, time_list):
        if c in notations:
            parts = notations[c].split("-")
            for p in parts:
                find_parent(p, time_list)
        else:
            time_list.append(c.split("_")[0])

        return time_list

    for cluster in clusters:
        times = list()
        times = find_parent(cluster, times)
        tt[cluster] = list(sorted(set(times)))

    print("Cluster Instants: \n", cluster_instants)
    print("len of tt: \n", len(tt))
    print("Time trace: \n", tt)

    for cluster, val in tt.items():
        v = dict.fromkeys(val, 1)
        keys = natsort.natsorted(v.keys())
        d_new = dict(OrderedDict((k, v[k]) for k in keys))
        tt[cluster] = d_new

    tt_df = pd.DataFrame.from_dict(tt, orient='index', dtype=object)
    tt_df = pd.DataFrame.from_dict(tt, orient='index', dtype=object)
    tt_df = tt_df.fillna(0)
    tt_df = tt_df.astype(object)

    styled = (tt_df.style.applymap(lambda v: 'background-color: %s' % 'grey' if v == 1 else 'red').
              applymap(lambda v: 'background-color: %s' % 'red' if v == 0 else 'grey'))
    styled.to_excel(tt_file, engine='openpyxl')
